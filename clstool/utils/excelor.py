import itertools
import pandas as pd
import numpy as np
import time
import os
import psutil
from openpyxl import load_workbook
from zipfile import BadZipFile

np.random.seed(42)

Tabel_Dict = {
    'Model_' : ['Model1', 'Model2', 'Model3', 'Model4', 'Model5', 'Model6'],
    'Dataset_' : ['Dataset1', 'Dataset2', 'Dataset3'],
    'Metric_' : ['Metric1', 'Metric2', 'Metric3', 'Metric4', 'Metric5'],
}

class Excelor:

    def __init__(self) -> None:
        self.df_m = None
        self.df_n = None
        self.df_info = {}
        
    def get_info(self, df_n):
        self.df_info = {}
        for i in range(len(df_n.columns)-1):
            self.df_info[df_n.columns[i]] = df_n.iloc[:, i].unique().tolist()

    def create(self, tabel_dict = Tabel_Dict):
        if self.df_m is None:
            index = pd.MultiIndex.from_product(list(tabel_dict.values()), names=list(tabel_dict.keys()))
            data = np.zeros(len(index))
            self.df_m = pd.DataFrame(data, index=index, columns=['Value'])
            self.df_n = self.df_m.reset_index()
            self.get_info(self.df_n)
        else:
            raise ValueError('Dataframe is already exists')
     
    def read_muti_index_xls(self, xls_dir):
        self.df_n = pd.read_excel(xls_dir)
        self.get_info(self.df_n)
        self.df_m = self.df_n.set_index(list(self.df_info.keys())).astype(float)
    
    def read_pivot_xls(self, xls_dir):
        self.df_n = pd.read_excel(xls_dir)
  
    def save_xls(self, xls_dir):
        self.df_n = self.df_m.reset_index()
        self.df_n.to_excel(xls_dir, index=False) 
        os.chmod(xls_dir, 0o777)
        
    def delete_xls(self, xls_dir):
        os.remove(xls_dir)

    def add_data(self, fixed_ids:list, score):
        tup = ()
        for key, value in self.df_info.items():
            loc = False
            for i in range(len(fixed_ids)):                
                if fixed_ids[i] in value:
                    tup = tup + (fixed_ids[i],)
                    fixed_ids.pop(i)
                    loc = True
                    break
            if loc == False:
                raise ValueError(f'No matching value in \'{key}\' was found from \'fixed_ids\'')
        self.df_m.loc[tup,'Value'] = round(score, 2)
    
    def custom_sort_key(self, x, tabel_dict, Level, Order):
        tabel_dicts = []
        levels = list(self.df_info.keys())
        for i in range(len(levels)):
            if i == levels.index(Level):
                tabel_dicts.append(Order.index(x[i]) if x[i] in Order else float('inf'))
            else:
                tabel_dicts.append(tabel_dict[levels[i]].index(x[i]))

        return tabel_dicts

    def insert_item(self, Level, Index, Order,value=0):
        if Index in self.df_info[Level]:
            raise ValueError(f'Insertion failure! Item \'{Index}\' is already existed in \'{Level}:{self.df_info[Level]}\'')

        levels = self.df_m.index.names
        new_indices = list(itertools.product(*[Order if level == Level else self.df_m.index.unique(level).tolist() for level in levels]))

        new_indices_ = []
        for i in new_indices:
            if i[levels.index(Level)] == Index:
                new_indices_.append(i)

        new_df = pd.DataFrame(index=pd.MultiIndex.from_tuples(new_indices_, names=levels))
        new_df['Value'] = value
        self.df_m = pd.concat([self.df_m, new_df])

        sorted_index = sorted(self.df_m.index, key=lambda x: self.custom_sort_key(x, tabel_dict, Level, Order))
        self.df_m = self.df_m.loc[sorted_index]

        self.df_n = self.df_m.reset_index()
        self.get_info(self.df_n)
    
    def delete_item(self, Level, Index):
        self.df_m = self.df_m.drop(index=Index, level=Level)
        self.df_n = self.df_m.reset_index()
        self.get_info(self.df_n)
        
    def pivot_table(self, fixed_ids:list, pxls_dir) :
        print(f"==>> len(list(self.df_info.keys())): {len(list(self.df_info.keys()))}")
        print(f"==>> len(fixed_ids): {len(fixed_ids)}")
        if len(list(self.df_info.keys()))-len(fixed_ids) != 2:
            raise ValueError('fixed_ids num error')
            
        indices_tup = ()
        indices_unfixed = [] 
        for key, value in self.df_info.items():
            flag = -1
            for i in range(len(fixed_ids)):                
                if fixed_ids[i] in value:
                    flag = i
            if flag==-1:
                indices_tup = indices_tup + (slice(None),) 
                indices_unfixed.append(key)
            else:
                indices_tup = indices_tup + (fixed_ids[flag],) 
        df_m_s = self.df_m.loc[indices_tup, :]
        print(f"==>> Pivot Items: {fixed_ids}")
        df_m_s_p = df_m_s.pivot_table(index=indices_unfixed[0], columns=indices_unfixed[1], values='Value', sort=False)
        df_m_s_p.reset_index().to_excel(pxls_dir, index=False) 
        os.chmod(pxls_dir, 0o777)
        return df_m_s_p
        
def number_to_letter(number):
    if 1 <= number <= 6:
        return chr(ord('A') + number - 1)
    elif 7 <= number <= 12:
        return chr(ord('a') + number - 7)
    else:
        return None      

def save_data_to_xls(fixed_ids:list, score, xls_dir):
    row, column = 0, 0
    wb = load_workbook(xls_dir)
    sheet = wb['Sheet1']
    for r in range(2, sheet.max_row + 1):  # row = 0  class_name
        value = []
        flag = 0
        for i in range(len(fixed_ids)):
            if sheet.cell(row=r, column=i+1).value == fixed_ids[i]:
                flag += 1
                continue
            else:
                break
        if flag == len(fixed_ids):
            row = r
            column = flag+1
            break
    letter = number_to_letter(column)
    loc = letter + str(row)
    score = round(score, 2)
    sheet[loc] = score
    wb.save(xls_dir)
    print(f'{fixed_ids}: {score} is saved')

def is_file_locked(filepath):
    """检查文件是否被锁定"""
    for proc in psutil.process_iter(['pid', 'open_files']):
        for file in proc.info['open_files'] or []:
            if file.path == filepath:
                return True
    return False

def load_excel_file_with_retry(file_path, retries=5, wait_time=5):
    """加载 Excel 文件并处理异常情况"""
    for attempt in range(retries):
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"The file {file_path} does not exist.")
            
            if is_file_locked(file_path):
                print(f"Attempt {attempt + 1}: File is locked. Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
                continue

            wb = load_workbook(file_path)
            return wb
        
        except (PermissionError, BadZipFile) as e:
            print(f"Attempt {attempt + 1}: {e}. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)

    raise Exception(f"Failed to open file after {retries} attempts.")

def save_datas_to_xls(fixed_ids:list, xls_dir):
    row, column = 0, 0
    wb = load_excel_file_with_retry(xls_dir)
    sheet = wb['Sheet1']

    for d in range(len(fixed_ids)):
        for r in range(2, sheet.max_row + 1):  # row = 0  class_name
            flag = 0
            for c in range(len(fixed_ids[d]) - 1):
                if sheet.cell(row=r, column=c+1).value == fixed_ids[d][c]:
                    flag += 1
                    continue
                else:
                    break
            if flag == (len(fixed_ids[d]) - 1):
                row = r
                column = flag + 1
                break
        letter = number_to_letter(column)
        loc = letter + str(row)
        score = round(fixed_ids[d][-1], 2)
        sheet[loc] = score
        sheet[loc].number_format = '0.00'

    wb.save(xls_dir)
    print(f'datas is saved to {xls_dir}')
        

if __name__ == '__main__':
    Fairxls = Excelor()

    xls_dir = './result/xls/Fairxls_corr.xlsx'
    # tabel_dict = {
    #     'Model': ["resnet18", "resnet34", "resnet50", "resnet101", "resnet152",
    #               "vgg11", "vgg13", "vgg16", "vgg19",
    #               "densenet121", "densenet169", "densenet201",
    #               "mobilenet_v2", "mobilenet_v3_small",
    #               "vit_small_patch32_224", "vit_base_patch32_224", "vit_large_patch32_224",
    #               "swin_tiny_patch4_window7_224", "swin_small_patch4_window7_224",
    #               "deit_tiny_patch16_224", "deit_small_patch16_224", "deit_base_patch16_224",
    #               "mixer_s16_224", "mixer_b16_224"],
    #     'Dataset': ['celeba', 'fonts-v1'],
    #     'Task': ['A','B','C'],
    #     'SA_Scale': ['small', 'medium', 'large'],
    #     'Metric': ['ACC', 'DP', 'EOpp', 'EOdd', 'Tol', 'Dev', 'Cou'],
    # }

    tabel_dict = {
        'Model': ["resnet50", "vit_small_patch32_224"],
        'Dataset': ['celeba'],
        'Main_attr': ['Attractive', 'Heavy_Makeup', 'Male', 'Young', 'Chubby', 'Smiling', 'Blurry',
            'High_Cheekbones', 'Narrow_Eyes', 'Rosy_Cheeks', 'Arched_Eyebrows', 'Big_Lips', 'Big_Nose', 'Bangs', 'Bags_Under_Eyes', 'Bald', 'Double_Chin', 'Goatee', 'Mustache', 'No_Beard'],
        'Sub_attr': ['Attractive', 'Heavy_Makeup', 'Male', 'Young', 'Chubby', 'Smiling', 'Blurry',
            'High_Cheekbones', 'Narrow_Eyes', 'Rosy_Cheeks', 'Arched_Eyebrows', 'Big_Lips', 'Big_Nose', 'Bangs', 'Bags_Under_Eyes', 'Bald', 'Double_Chin', 'Goatee', 'Mustache', 'No_Beard'],
        'Metric': ['ACC', 'DP', 'EOpp', 'EOdd'],
    }

    # Fairxls.delete_xls(xls_dir)
    
    # Fairxls.create(tabel_dict)
    # print('==>> Fairxls.df_n: ', Fairxls.df_n)
    # print('==>> Fairxls.df_info:\n ', Fairxls.df_info)
    # Fairxls.save_xls(xls_dir)
    

    Fairxls.read_muti_index_xls(xls_dir)
    print('==>> Fairxls.df_info:\n',Fairxls.df_info)

    
    # Fairxls.add_data(["resnet18", 'A', 'celeba', 'small', 'DP'], 1.274)
    # Fairxls.save_xls(xls_dir)


    # Fairxls.insert_item('Model', 'vgg19', ['vgg19', "resnet18", "vgg11", 'vgg13', "resnet34", "resnet50"])
    # Fairxls.save_xls(xls_dir)
    # print('==>> Fairxls.df_info:\n',Fairxls.df_info)


    # Fairxls.delete_item('Model', 'vgg20')
    # Fairxls.save_xls(xls_dir)
    # print('==>> Fairxls.df_info:\n',Fairxls.df_info)

    pxls_dir = './result/xls/Pivotxls_corr.xlsx'
    # fixed_ids = ['celeba','A','medium']
    fixed_ids = ['resnet50','celeba','EOdd']
    pivot_table = Fairxls.pivot_table(fixed_ids, pxls_dir)


    # fixed_ids = ["resnet18", 'celeba', 'A', 'small', 'Cou']
    # # fixed_ids = ["mixer_b16_224", 'fonts-v1', 'C', 'large', 'ACC']
    # score = 0.356
    # save_data_to_xls(fixed_ids, score, xls_dir)


    # saved_datas = [["resnet18", 'celeba', 'A', 'small', 'ACC', 1.99],
    #              ["resnet18", 'celeba', 'A', 'small', 'DP', 2.23],
    #              ["resnet18", 'celeba', 'A', 'small', 'EOpp', 3.23],
    #              ["resnet18", 'celeba', 'A', 'small', 'EOdd', 4.23],
    #              ["resnet18", 'celeba', 'A', 'small', 'Tol', 5.23],
    #              ["resnet18", 'celeba', 'A', 'small', 'Dev', 6.23],
    #              ["resnet18", 'celeba', 'A', 'small', 'Cou', 7.23]]
    # save_datas_to_xls(saved_datas, xls_dir)
    
    



    
    
